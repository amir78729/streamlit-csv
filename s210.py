import pandas as pd
import streamlit as st
import io
import os
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# Constants
SEP = ';'
ENCODING = 'ISO-8859-1'

SHEET_MAPPING_HIGH_INCOME = {
    "FörlSj>2": "Sjuk Fortsättningsnivå>År",
    "Sjuk": "Sjuk",
    "Sj>2år": "Sjuk >år",
    "Sjukförl": "Sjuk Fortsättningsnivå",
    "Arbskliv": "Arbetsskadelivränta",
    "Tsjer": "Sjukers",
    "Tsjer>2": "Sjukers",
    "Arbfl>2": "Sjuk Fortsättningsnivå>År",
    "ReFör>år": "Sjuk Fortsättningsnivå>År",
    "FörSj>hö": "Sjuk Fortsättningsnivå>År",
    "FörSj>2": "Sjuk Fortsättningsnivå>År",
    "SjNysjpe": "Sjuk",
    "SjukAmos": "Sjuk",
    "SjFörbm5": "Sjuk",
    "Läkarbesök": "Sjuk",
    "SjukFK": "Sjuk",
    "Arbskada": "Sjuk",
    "Rehab>2": "Sjuk >år",
    "SjNyejse": "Sjuk >år",
    "SjEjsemg": "Sjuk >år",
    "Sjförlhö": "Sjuk Fortsättningsnivå",
    "Arbsförl": "Sjuk Fortsättningsnivå"
}

SHEET_MAPPING_LOW_INCOME = {
    "Arbsk>2": "Sjuk >år",
    "FörlSj>2": "Sjuk Fortsättningsnivå>År",
    "Sj>180": "Sjuk >år",
    "Sj>180Se": "Sjuk",
    "Sj>2år": "Sjuk >år",
    "SjFK>år": "Sjuk >år",
    "Sjuk": "Sjuk",
    "Arbskliv": "Arbetsskadelivränta",
    "Tsjer": "Sjukers",
    "Tsjer>2": "Sjukers",
    "Arbfl>2": "Sjuk Fortsättningsnivå>År",
    "ReFör>år": "Sjuk Fortsättningsnivå>År",
    "FörSj>hö": "Sjuk Fortsättningsnivå>År",
    "FörSj>2": "Sjuk Fortsättningsnivå>År",
    "SjNysjpe": "Sjuk",
    "SjukAmos": "Sjuk",
    "SjFörbm5": "Sjuk",
    "Läkarbesök": "Sjuk",
    "SjukFK": "Sjuk",
    "Arbskada": "Sjuk",
    "Rehab>2": "Sjuk >år",
    "SjNyejse": "Sjuk >år",
    "SjEjsemg": "Sjuk >år",
    "Sjförlhö": "Sjuk Fortsättningsnivå",
    "Arbsförl": "Sjuk Fortsättningsnivå",
}

SHEET_NAMES = [
    "Sjuk",
    "Sjuk >år",
    "Sjuk Fortsättningsnivå",
    "Sjuk Fortsättningsnivå>År",
    "Sjukers",
    "Arbetsskadelivränta",
]

# Styling
header_font = Font(bold=True)
bold_font = Font(bold=True)
center_alignment = Alignment(horizontal='center')
border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
purple_font = Font(color="800080", bold=True)
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

# Functions
def apply_cell_style(cell, col_name, value):
    cell.border = border_style
    if col_name == "Verklig lön":
        try:
            number_value = int(str(value).split(',')[0])
            cell.alignment = center_alignment
            cell.font = purple_font if number_value >= 49000 else bold_font
            cell.value = number_value
            cell.number_format = '#,##0'
        except:
            pass
    elif col_name == "Sem Grp" and str(value) in ['8', '9']:
        cell.fill = orange_fill

def copy_rows_to_sheet(sheet, headers, rows, start_row=2):
    for r_offset, row in enumerate(rows, start=start_row):
        for c_idx, value in enumerate(row, 1):
            col_name = headers[c_idx - 1]
            cell = sheet.cell(row=r_offset, column=c_idx, value=value)
            apply_cell_style(cell, col_name, value)

def extract_matching_rows(df, headers):
    matched = {sheet: [] for sheet in SHEET_NAMES}
    
    for row in dataframe_to_rows(df, index=False, header=True):
        if row == headers:
            continue
        row_dict = dict(zip(headers, row))
        try:
            income = int(str(row_dict.get("Verklig lön", "0")).split(',')[0])
            if income >= 49000:
                frn_ors_value = row_dict.get("Frn Ors", "").strip()
                if frn_ors_value in SHEET_MAPPING_HIGH_INCOME:
                    st.write('- 🟣 "Frn Ors" value = `{}` and Income >= 49000: Moving to `{}` sheet...'.format(frn_ors_value, SHEET_MAPPING_HIGH_INCOME[frn_ors_value]))
                    matched[SHEET_MAPPING_HIGH_INCOME[frn_ors_value]].append(row)
                    
            else:
                frn_ors_value = row_dict.get("Frn Ors", "").strip()
                if frn_ors_value in SHEET_MAPPING_LOW_INCOME:
                    st.write('- ⚫ "Frn Ors" value = `{}` and Income < 49000: Moving to `{}` sheet...'.format(frn_ors_value, SHEET_MAPPING_LOW_INCOME[frn_ors_value]))
                    matched[SHEET_MAPPING_LOW_INCOME[frn_ors_value]].append(row)
        except:
            pass
    return matched

# Streamlit UI
st.title("S2_10: Sjuk mer än 365 dagar")
keep_all_data = st.checkbox("Keep original data as a worksheet", value=False)
uploaded_file = st.file_uploader("Upload your file", type=["csv", "xlsx"])
st.info("Don't worry! All calculations are local and data is not shared.")

if uploaded_file is not None:
    base_filename = os.path.splitext(uploaded_file.name)[0]
    excel_result_filename = f"{base_filename}-result.xlsx"

    file_extension = os.path.splitext(uploaded_file.name)[1].lower()
    with st.status("Reading File"):
        if file_extension == '.csv':
            df = pd.read_csv(uploaded_file, sep=SEP, encoding=ENCODING)
        elif file_extension == '.xlsx':
            df = pd.read_excel(uploaded_file)
        st.write(df)

    wb = Workbook()

    if keep_all_data:
        ws_all = wb.active
        ws_all.title = "All data"
    else:
        wb.remove(wb.active)
        ws_all = None  

    sheets = {name: wb.create_sheet(title=name) for name in SHEET_NAMES}
    wb.active = wb[sheets["Sjuk"].title]

    with st.status("Updating cell values"):
        headers = df.columns.tolist()

        if keep_all_data:
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    col_name = headers[c_idx - 1]
                    cell = ws_all.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = header_font
                    else:
                        apply_cell_style(cell, col_name, value)

        # Apply headers
        for sheet in sheets.values():
            for c_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=c_idx, value=header)
                cell.font = header_font
                cell.border = border_style

        # Extract and copy rows based on the mapping
        matched_rows = extract_matching_rows(df, headers)
        for sheet_name, rows in matched_rows.items():
            copy_rows_to_sheet(sheets[sheet_name], headers, rows, start_row=2)

    # Apply auto-filters
    if ws_all:
        ws_all.auto_filter.ref = ws_all.dimensions
    for sheet in sheets.values():
        sheet.auto_filter.ref = sheet.dimensions

    with st.status("Creating the Excel file"):
        for sheet in sheets:
            st.write("`{}` sheet:".format(sheet))
            st.write(pd.DataFrame(matched_rows[sheet]))
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

    st.download_button(
        label="Download Excel",
        data=output,
        file_name=excel_result_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
