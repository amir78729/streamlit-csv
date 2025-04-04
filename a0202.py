import pandas as pd
import streamlit as st
import io
import os
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from datetime import datetime
import calendar

SEP = ';'
ENCODING = 'ISO-8859-1'


# Styling
header_font = Font(bold=True)
border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Streamlit UI
st.title("A02_02: Sjuk o frånvaro med omf")
uploaded_file = st.file_uploader("Upload your file", type=["csv", "xlsx"])



is_before_payment = st.checkbox("Is this process before payment?", value=True)

month_numbers = [str(i).zfill(2) for i in range(1, 13)]

# Use the selectbox with format_func to display month names alongside the month number
month = st.selectbox(
    "Select Month",
    month_numbers,
    format_func=lambda x: f"{calendar.month_name[int(x)]} ({x})",  # Display month name and number (e.g., January (01))
)
st.info("Don't worry! All calculations are local and data is not shared.")

if uploaded_file is not None:
    base_filename = os.path.splitext(uploaded_file.name)[0]
    excel_result_filename = f"{base_filename}-result.xlsx"
    # Detect file extension to handle accordingly
    file_extension = os.path.splitext(uploaded_file.name)[1].lower()

    with st.status("Reading File"):
        if file_extension == '.csv':
            df = pd.read_csv(uploaded_file, sep=SEP, encoding=ENCODING)
        elif file_extension == '.xlsx':
            df = pd.read_excel(uploaded_file)

        df.columns = df.columns.str.strip()
        st.write(df)
        df['Sem Gfom'] = df['Sem Gfom'].astype(str)

    # Filter data
    df_filtered = pd.DataFrame()

    with st.status("Step 1: Filtering data"):
        df_filtered = df[
            (df['Lbertom'] == 0)  # Check Lbertom value
            & ((df['Semester'] == 'Semsjuk') | (df['Semester'] == 'Semsjuk1'))  # Check if Semester is 'Semsjuk' or 'Semsjuk1'
            & (df['Sjuk Korr'] == 1)  # Check Sjuk Korr value
            & ((df['Sem Omf'] + df['Annan Omf']) == 1)  # Sum of 'Sem Omf' and 'Annan Omf' equals 1
            | ((df['Sem Omf'] + df['Annan Omf']) != 1) & (df['Sem Korr'] == 2)  # OR condition for Sem Korr == 2
        ]

        # Filter by month
        df_filtered = df_filtered[df_filtered['Sem Gfom'].str[:6] == f"{datetime.now().year}{month}"]

        st.write('Rows that match filtering criteria:')
        st.write(df_filtered)

        # Merge to remove filtered data from df
        try:
            df = pd.merge(df, df_filtered, how="outer", indicator=True).query("_merge != 'both'").drop('_merge', axis=1).reset_index(drop=True)
        except: 
            pass

        st.write('Filtered data removed, resulting data:')
        st.write(df)

    if is_before_payment:
        with st.status("Filter for before payment"):
            df_filtered = df[((df['Sem Omf'] + df['Annan Omf']) == 1)]

            # Filter by month
            df_filtered = df_filtered[df_filtered['Sem Gfom'].str[:6] == f"{datetime.now().year}{month}"]

            st.write('Rows that match filtering criteria:')
            st.write(df_filtered)

            # Merge to remove filtered data from df
            try:
                df = pd.merge(df, df_filtered, how="outer", indicator=True).query("_merge != 'both'").drop('_merge', axis=1).reset_index(drop=True)
            except: 
                pass

            st.write('Filtered data removed, resulting data:')
            st.write(df)

    with st.status("Filter based on Semsjuk and Semsjuk1 Semesters"):
        df_filtered = df[
            ((df['Semester'] == 'Semsjuk') | (df['Semester'] == 'Semsjuk1'))  # Check if Semester is 'Semsjuk' or 'Semsjuk1'
            & ((df['Sem Omf'] + df['Annan Omf']) == 1)  # Sum of 'Sem Omf' and 'Annan Omf' equals 1
        ]

        # Filter by month
        df_filtered = df_filtered[df_filtered['Sem Gfom'].str[:6] == f"{datetime.now().year}{month}"]

        st.write('Rows that match filtering criteria:')
        st.write(df_filtered)

        # Merge to remove filtered data from df
        try:
            df = pd.merge(df, df_filtered, how="outer", indicator=True).query("_merge != 'both'").drop('_merge', axis=1).reset_index(drop=True)
        except: 
            pass

        st.write('Filtered data removed, resulting data:')
        st.write(df)

    with st.status("Filter based on Sembet and Sembet1 Semesters"):
        df_filtered = df[
            ((df['Semester'] == 'Sembet') | (df['Semester'] == 'Sembet1'))  # Check if Semester is 'Semsjuk' or 'Semsjuk1'
            & ((df['Sjuk Korr']) == 1)
            & ((df['Sem Korr']) == 2)
        ]

        # Filter by month
        df_filtered = df_filtered[df_filtered['Sem Gfom'].str[:6] == f"{datetime.now().year}{month}"]

        st.write('Rows that match filtering criteria:')
        st.write(df_filtered)

        # Merge to remove filtered data from df
        try:
            df = pd.merge(df, df_filtered, how="outer", indicator=True).query("_merge != 'both'").drop('_merge', axis=1).reset_index(drop=True)
        except: 
            pass

        st.write('Filtered data removed, resulting data:')
        st.write(df)
        
        
    with st.status("Filter Sjuk Korr = 2"):
        df['Pnr'] = df['Pnr'].astype(str).str.strip()

        # After filtering by 'Sjuk Korr' == 1
        df_filtered = df[df['Sjuk Korr'] == 1]

        # Reset index to avoid issues with groupby
        df_filtered = df_filtered.reset_index(drop=True)

        # Count occurrences of each 'Pnr'
        pnr_counts = df_filtered.groupby('Pnr').size()

        # Filter out rows where 'Pnr' appears only once (unique Pnr)
        df_filtered = df_filtered[df_filtered['Pnr'].isin(pnr_counts[pnr_counts > 1].index)]

        st.write('Rows with unique Pnr removed:')
        st.write(df_filtered)

        # Merge to remove filtered data from df
        try:
            df = pd.merge(df, df_filtered, how="outer", indicator=True).query("_merge != 'both'").drop('_merge', axis=1).reset_index(drop=True)
        except: 
            pass

        st.write('Filtered data removed, resulting data:')
        st.write(df)
        
        

    with st.status("Filter based on Förv Bolag value"):
        df_filtered = df[
            ((df['Förv Bolag'] == 'Arbvux') | (df['Förv Bolag'] == 'AMA'))
        ]

        # Filter by month
        df_filtered = df_filtered[df_filtered['Sem Gfom'].str[:6] == f"{datetime.now().year}{month}"]

        st.write('Rows that match filtering criteria:')
        st.write(df_filtered)

        # Merge to remove filtered data from df
        try:
            df = pd.merge(df, df_filtered, how="outer", indicator=True).query("_merge != 'both'").drop('_merge', axis=1).reset_index(drop=True)
        except: 
            pass

        st.write('Filtered data removed, resulting data:')
        st.write(df)
        
        
    with st.status("Make cells yellow"):
      # Filter the rows that should be highlighted in yellow
      df_yellow = df[
          (
              (df['Lbertom'] == 0)  # Check Lbertom value
              & ((df['Semester'] == 'Semsjuk') | (df['Semester'] == 'Semsjuk1'))  # Check if Semester is 'Semsjuk' or 'Semsjuk1'
              & (df['Sjuk Korr'] == 1)  # Check Sjuk Korr value
              & ((df['Sem Omf'] + df['Annan Omf']) != 1)  # Sum of 'Sem Omf' and 'Annan Omf' equals 1
          ) | (
              is_before_payment
              & ((df['Sem Omf'] + df['Annan Omf']) != 1)  # Sum of 'Sem Omf' and 'Annan Omf' equals 1
          ) | (
              ((df['Semester'] == 'Semsjuk') | (df['Semester'] == 'Semsjuk1'))
              & ((df['Sem Omf'] + df['Annan Omf']) != 1)  # Sum of 'Sem Omf' and 'Annan Omf' equals 1
          )
      ]
      
      # Filter by month
      df_yellow = df_yellow[df_yellow['Sem Gfom'].str[:6] == f"{datetime.now().year}{month}"]

      st.write("Filtered data that should be highlighted in yellow:")
      st.write(df_yellow)

    
    with st.status("Creating Excel File"):
            # Now, we create the Excel file and apply the yellow fill
      output = io.BytesIO()

      # Create an Excel file using openpyxl
      wb = Workbook()
      ws_all = wb.active
      ws_all.title = "Filtered Data"

      # Write the data to the Excel file and apply the yellow fill based on conditions
      headers = df.columns.tolist()

      for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
          for c_idx, value in enumerate(row, 1):
              col_name = headers[c_idx - 1]
              cell = ws_all.cell(row=r_idx, column=c_idx, value=value)

              # Apply formatting for the header row
              if r_idx == 1:
                  cell.font = header_font
              else:
                  cell.border = border_style

                  # Apply yellow fill for 'Sem Omf' or 'Annan Omf' columns based on specific condition
                  if col_name == 'Sem Omf' or col_name == 'Annan Omf':
                      # Apply the yellow fill only if the row meets the conditions in df_yellow
                      condition = (
                          (df['Lbertom'] == 0)  # Check Lbertom value
                          & ((df['Semester'] == 'Semsjuk') | (df['Semester'] == 'Semsjuk1'))  # Check if Semester is 'Semsjuk' or 'Semsjuk1'
                          & (df['Sjuk Korr'] == 1)  # Check Sjuk Korr value
                          & ((df['Sem Omf'] + df['Annan Omf']) != 1)  # Sum of 'Sem Omf' and 'Annan Omf' equals 1
                      ) | (
                          is_before_payment
                          & ((df['Sem Omf'] + df['Annan Omf']) != 1)  # Sum of 'Sem Omf' and 'Annan Omf' equals 1
                      ) | (
                          ((df['Semester'] == 'Semsjuk') | (df['Semester'] == 'Semsjuk1'))
                          & ((df['Sem Omf'] + df['Annan Omf']) != 1)  # Sum of 'Sem Omf' and 'Annan Omf' equals 1
                      )

                      # Check if the row meets the condition
                      if condition.iloc[r_idx - 2]:  # Adjust index for 0-based indexing in pandas
                          cell.fill = yellow_fill
      # Apply auto-filters
      ws_all.auto_filter.ref = ws_all.dimensions

      # Save the workbook to the output stream
      wb.save(output)
      processed_data = output.getvalue()

      # Provide download button
    st.download_button(
      label="Download Excel",
      data=processed_data,
      file_name=excel_result_filename,
      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
  )
